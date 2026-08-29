import io
import os
import zipfile
import json
import xml.etree.ElementTree as ET
import pyproj
import simplekml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from shapely.geometry import Polygon, Point, shape
from shapely.ops import transform
from flask import Flask, request, jsonify, send_file, send_from_directory

app = Flask(__name__)

# ==========================================
# 🗺️ ระบบค้นหาสถานที่ (Local Reverse Geocoding) ทำงานหลังบ้าน
# ==========================================
tambon_features = []
geojson_path = 'tambon_thailand.json'

# โหลดข้อมูลเข้าหน่วยความจำแค่ครั้งเดียวตอนเปิดเซิร์ฟเวอร์
if os.path.exists(geojson_path):
    print("กำลังโหลดข้อมูลขอบเขตตำบลเพื่อใช้ค้นหาสถานที่ (ทำงานเบื้องหลัง)...")
    try:
        with open(geojson_path, 'r', encoding='utf-8') as f:
            gj_data = json.load(f)
            for feat in gj_data.get('features', []):
                if feat.get('geometry'):
                    geom = shape(feat['geometry'])
                    tambon_features.append({
                        'geom': geom,
                        'bounds': geom.bounds,
                        'props': feat.get('properties', {})
                    })
        print(f"โหลดข้อมูลตำบลสำเร็จ {len(tambon_features)} รายการ พร้อมใช้งาน!")
    except Exception as e:
        print(f"⚠️ ไม่สามารถโหลดไฟล์ {geojson_path} ได้: {e}")
else:
    print(f"⚠️ ไม่พบไฟล์ {geojson_path} ระบบจะไม่สามารถดึงชื่อสถานที่อัตโนมัติได้")

def get_local_location(lon, lat):
    """ฟังก์ชันเช็ค Point in Polygon แบบเร็ว"""
    if not tambon_features:
        return ""
    
    pt = Point(lon, lat)
    for item in tambon_features:
        minx, miny, maxx, maxy = item['bounds']
        # 1. เช็ค Bounding Box ก่อน (ไวมาก)
        if minx <= lon <= maxx and miny <= lat <= maxy:
            # 2. เช็ค Polygon ละเอียด (ถ้าผ่านเงื่อนไขแรก)
            if item['geom'].contains(pt):
                props = item['props']
                # ดึงฟิลด์ตามตาราง Attribute ของผู้การ
                tam = props.get('TAM_NAM_T', '')
                amp = props.get('AMPHOE_T', '')
                prov = props.get('PROV_NAM_T', '')
                
                parts = []
                if tam: parts.append(tam)
                if amp: parts.append(amp)
                if prov: parts.append(prov)
                
                return " ".join(parts).strip()
    return ""


def dd_to_dms(dd, is_lat=True):
    direction = ("N" if dd >= 0 else "S") if is_lat else ("E" if dd >= 0 else "W")
    dd = abs(dd)
    degrees = int(dd)
    minutes_float = (dd - degrees) * 60
    minutes = int(minutes_float)
    seconds = round((minutes_float - minutes) * 60, 2)
    return f'{degrees}° {minutes}\' {seconds:.2f}" {direction}'

@app.route('/')
def index():
    if os.path.exists('index.html'):
        return send_from_directory('.', 'index.html')
    elif os.path.exists('templates_index.html'):
        return send_from_directory('.', 'templates_index.html')
    return "<h3>⚠️ หาไฟล์ HTML ไม่เจอ</h3>", 404

@app.route('/<path:filename>')
def serve_static(filename):
    if os.path.exists(filename):
        return send_from_directory('.', filename)
    return jsonify({"error": "File not found"}), 404

# ==========================================
# ✈️ AIRPLANE API
# ==========================================
@app.route('/api/calculate', methods=['POST'])
def calculate_area():
    data = request.json or {}
    buf_nw = data.get('buffer_nw', [0, 0])
    buf_ne = data.get('buffer_ne', [0, 0])
    buf_se = data.get('buffer_se', [0, 0])
    buf_sw = data.get('buffer_sw', [0, 0])

    corners = {
        "NW": {"lat": buf_nw[0], "lng": buf_nw[1], "lat_dms": dd_to_dms(buf_nw[0], True), "lng_dms": dd_to_dms(buf_nw[1], False)},
        "NE": {"lat": buf_ne[0], "lng": buf_ne[1], "lat_dms": dd_to_dms(buf_ne[0], True), "lng_dms": dd_to_dms(buf_ne[1], False)},
        "SE": {"lat": buf_se[0], "lng": buf_se[1], "lat_dms": dd_to_dms(buf_se[0], True), "lng_dms": dd_to_dms(buf_se[1], False)},
        "SW": {"lat": buf_sw[0], "lng": buf_sw[1], "lat_dms": dd_to_dms(buf_sw[0], True), "lng_dms": dd_to_dms(buf_sw[1], False)}
    }
    return jsonify({"status": "success", "corners": corners})

@app.route('/api/download', methods=['POST'])
def download_package():
    data = request.json or {}
    project_name = data.get('project_name', 'NOTAM_PROJECT').strip().replace(' ', '_')
    sheets = data.get('sheets', [])

    in_nw = data.get('inner_nw', [0, 0])
    in_ne = data.get('inner_ne', [0, 0])
    in_se = data.get('inner_se', [0, 0])
    in_sw = data.get('inner_sw', [0, 0])

    buf_nw = data.get('buffer_nw', [0, 0])
    buf_ne = data.get('buffer_ne', [0, 0])
    buf_se = data.get('buffer_se', [0, 0])
    buf_sw = data.get('buffer_sw', [0, 0])

    merged_sheet_name = "_".join([s.replace(" ", "") for s in sheets]) if sheets else "Merged_Block"

    kml_block = simplekml.Kml()
    pol_block = kml_block.newpolygon(name=merged_sheet_name)
    pol_block.outerboundaryis = [(in_sw[1], in_sw[0]), (in_se[1], in_se[0]), (in_ne[1], in_ne[0]), (in_nw[1], in_nw[0]), (in_sw[1], in_sw[0])]
    pol_block.style.polystyle.color = '4000ffff'
    pol_block.style.linestyle.color = 'ff00ffff'
    pol_block.style.linestyle.width = 3

    kml_notam = simplekml.Kml()
    pol_notam = kml_notam.newpolygon(name=f"{project_name}_Buffer")
    pol_notam.outerboundaryis = [(buf_sw[1], buf_sw[0]), (buf_se[1], buf_se[0]), (buf_ne[1], buf_ne[0]), (buf_nw[1], buf_nw[0]), (buf_sw[1], buf_sw[0])]
    pol_notam.style.polystyle.color = '400000ff'
    pol_notam.style.linestyle.color = 'ff0000ff'
    pol_notam.style.linestyle.width = 2

    for name, p in [("NW", buf_nw), ("NE", buf_ne), ("SE", buf_se), ("SW", buf_sw)]:
        pnt = kml_notam.newpoint(name=name, coords=[(p[1], p[0])])
        pnt.description = f"Lat: {p[0]:.6f}, Lon: {p[1]:.6f}\nDMS: {dd_to_dms(p[0], True)}, {dd_to_dms(p[1], False)}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coordinates"
    ws.views.sheetView[0].showGridLines = True
    ws.append(["Corner Position", "Latitude (DMS)", "Longitude (DMS)", "Latitude (DD)", "Longitude (DD)"])

    rows = [
        ["NW (บน-ซ้าย)", dd_to_dms(buf_nw[0], True), dd_to_dms(buf_nw[1], False), round(buf_nw[0], 6), round(buf_nw[1], 6)],
        ["NE (บน-ขวา)", dd_to_dms(buf_ne[0], True), dd_to_dms(buf_ne[1], False), round(buf_ne[0], 6), round(buf_ne[1], 6)],
        ["SE (ล่าง-ขวา)", dd_to_dms(buf_se[0], True), dd_to_dms(buf_se[1], False), round(buf_se[0], 6), round(buf_se[1], 6)],
        ["SW (ล่าง-ซ้าย)", dd_to_dms(buf_sw[0], True), dd_to_dms(buf_sw[1], False), round(buf_sw[0], 6), round(buf_sw[1], 6)]
    ]
    for r in rows: ws.append(r)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    for col in range(1, 6):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")

    for r in range(2, 6):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col, width in {'A': 18, 'B': 22, 'C': 22, 'D': 18, 'E': 18}.items():
        ws.column_dimensions[col].width = width

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"{merged_sheet_name}.kml", kml_block.kml().encode('utf-8'))
        z.writestr(f"{project_name}_Buffer.kml", kml_notam.kml().encode('utf-8'))
        z.writestr(f"{project_name}_coordinates.xlsx", excel_bytes.getvalue())

    zip_buffer.seek(0)
    
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=f'{project_name}_Package.zip')


# ==========================================
# 🚁 UAV API
# ==========================================
@app.route('/api/calculate_uav', methods=['POST'])
def calculate_uav():
    data = request.get_json() or {}
    coords = data.get('coordinates', [])
    buffer_meters = float(data.get('buffer_meters', 50))
    
    if len(coords) < 3:
        return jsonify({"error": "Need at least 3 coordinates"}), 400
    
    poly = Polygon(coords)
    
    centroid_lon = poly.centroid.x
    centroid_lat = poly.centroid.y
    
    utm_zone = int((centroid_lon + 180) / 6) + 1
    epsg_code = f"326{utm_zone}" if centroid_lat >= 0 else f"327{utm_zone}"
    
    project_to_utm = pyproj.Transformer.from_crs('EPSG:4326', f'EPSG:{epsg_code}', always_xy=True).transform
    project_to_wgs84 = pyproj.Transformer.from_crs(f'EPSG:{epsg_code}', 'EPSG:4326', always_xy=True).transform
    
    poly_utm = transform(project_to_utm, poly)
    buffer_utm = poly_utm.buffer(buffer_meters, join_style=2)
    
    poly_wgs84 = transform(project_to_wgs84, poly_utm)
    buffer_wgs84 = transform(project_to_wgs84, buffer_utm)

    # วิ่งไปค้นหาชื่อสถานที่จากฟังก์ชัน Local ที่เราเขียนไว้ด้านบน
    location_name = get_local_location(centroid_lon, centroid_lat)
    
    return jsonify({
        "inner_coords": list(poly_wgs84.exterior.coords),
        "buffer_coords": list(buffer_wgs84.exterior.coords),
        "location_name": location_name # ส่งกลับไปให้หน้าเว็บพิมพ์ลงใน Textbox 
    })

@app.route('/api/upload_kml', methods=['POST'])
def upload_kml():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    filename = file.filename.lower()
    content = file.read()
    coords = []
    
    try:
        if filename.endswith('.geojson'):
            data = json.loads(content)
            for feat in data.get('features', []):
                geom = feat.get('geometry', {})
                if geom.get('type') == 'Polygon':
                    coords = geom.get('coordinates', [[]])[0]
                    break
                elif geom.get('type') == 'MultiPolygon':
                    coords = geom.get('coordinates', [[[]]])[0][0]
                    break
        elif filename.endswith('.kml'):
            root = ET.fromstring(content)
            for elem in root.iter():
                if elem.tag.endswith('coordinates'):
                    text = elem.text.strip()
                    pts = []
                    for part in text.split():
                        subparts = part.split(',')
                        if len(subparts) >= 2:
                            pts.append([float(subparts[0]), float(subparts[1])])
                    if len(pts) >= 3:
                        coords = pts
                        break
    except Exception as e:
        return jsonify({"error": f"Parsing error: {str(e)}"}), 400
        
    if not coords:
        return jsonify({"error": "Could not extract polygon coordinates from file"}), 400
        
    return jsonify({"status": "success", "coordinates": coords})

@app.route('/api/download_uav', methods=['POST'])
def download_uav():
    data = request.get_json() or {}
    project_name = data.get('project_name', 'UAV_PROJECT').strip().replace(' ', '_')
    location_name = data.get('location_name', '-') # รับค่า location มาลง Excel
    coords = data.get('coordinates', [])
    buffer_meters = float(data.get('buffer_meters', 50))
    
    if len(coords) < 3:
        return jsonify({"error": "Invalid coordinates"}), 400
        
    poly = Polygon(coords)
    centroid_lon = poly.centroid.x
    centroid_lat = poly.centroid.y
    
    utm_zone = int((centroid_lon + 180) / 6) + 1
    epsg_code = f"326{utm_zone}" if poly.centroid.y >= 0 else f"327{utm_zone}"
    
    project_to_utm = pyproj.Transformer.from_crs('EPSG:4326', f'EPSG:{epsg_code}', always_xy=True).transform
    project_to_wgs84 = pyproj.Transformer.from_crs(f'EPSG:{epsg_code}', 'EPSG:4326', always_xy=True).transform
    
    poly_utm = transform(project_to_utm, poly)
    buffer_utm = poly_utm.buffer(buffer_meters, join_style=2)
    
    poly_wgs84 = transform(project_to_wgs84, poly_utm)
    buffer_wgs84 = transform(project_to_wgs84, buffer_utm)
    
    kml_inner = simplekml.Kml()
    pol_inner = kml_inner.newpolygon(name=f"{project_name}_Mission_Area")
    pol_inner.outerboundaryis = [(c[0], c[1]) for c in poly_wgs84.exterior.coords]
    pol_inner.style.polystyle.color = '4000a5ff'
    pol_inner.style.linestyle.color = 'ff00a5ff'
    pol_inner.style.linestyle.width = 3

    kml_buffer = simplekml.Kml()
    pol_buf = kml_buffer.newpolygon(name=f"{project_name}_Buffer_{buffer_meters}m")
    buffer_coords = list(buffer_wgs84.exterior.coords)
    pol_buf.outerboundaryis = [(c[0], c[1]) for c in buffer_coords]
    pol_buf.style.polystyle.color = '400000ef'
    pol_buf.style.linestyle.color = 'ff0000ef'
    pol_buf.style.linestyle.width = 2

    for idx, pt in enumerate(buffer_coords[:-1], start=1):
        lon, lat = pt[0], pt[1]
        pnt = kml_buffer.newpoint(name=f"Buf_Pt_{idx}", coords=[(lon, lat)])
        pnt.description = f"Lat: {lat:.6f}, Lon: {lon:.6f}\nDMS: {dd_to_dms(lat, True)}, {dd_to_dms(lon, False)}"

    # ---------------------------------------------
    # 📊 EXCEL GENERATION (อัปเกรดใหม่)
    # ---------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mission_Data"
    ws.views.sheetView[0].showGridLines = False 
    
    # ส่วนหัวข้อมูล (Summary Section)
    ws.append(["📌 Project Name:", project_name])
    ws.append(["📍 Location:", location_name])
    ws.append(["🎯 Centroid Latitude:", dd_to_dms(centroid_lat, True), round(centroid_lat, 6)])
    ws.append(["🎯 Centroid Longitude:", dd_to_dms(centroid_lon, False), round(centroid_lon, 6)])
    ws.append(["📏 Buffer Distance:", f"{buffer_meters} Meters"])
    ws.append([]) 

    bold_font = Font(bold=True)
    for r in range(1, 6):
        ws.cell(row=r, column=1).font = bold_font

    table_start_row = 7
    ws.append(["Buffer Vertex", "Latitude (DMS)", "Longitude (DMS)", "Latitude (DD)", "Longitude (DD)"])

    for idx, pt in enumerate(buffer_coords[:-1], start=1):
        lon, lat = pt[0], pt[1]
        ws.append([f"Point {idx}", dd_to_dms(lat, True), dd_to_dms(lon, False), round(lat, 6), round(lon, 6)])

    header_fill = PatternFill(start_color="D96B27", end_color="D96B27", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    for col in range(1, 6):
        c = ws.cell(row=table_start_row, column=col)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")

    for r in range(table_start_row + 1, table_start_row + len(buffer_coords)):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col, width in {'A': 20, 'B': 22, 'C': 22, 'D': 18, 'E': 18}.items():
        ws.column_dimensions[col].width = width

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"{project_name}_Mission_Area.kml", kml_inner.kml().encode('utf-8'))
        z.writestr(f"{project_name}_Buffer_{buffer_meters}m.kml", kml_buffer.kml().encode('utf-8'))
        z.writestr(f"{project_name}_coordinates.xlsx", excel_bytes.getvalue())

    zip_buffer.seek(0)
    
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=f'{project_name}_UAV_Package.zip')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)
